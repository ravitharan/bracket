#! /usr/bin/env python3

import sys
import openpyxl
from itertools import combinations
import random
import argparse

ATTENDANCE_SHEET = "attendance"

SCHEDULE_SHEET = "quali_schedule"
WINNER_POINTS = 1

SCORE_SHEET = "quali_score"

LEFT_SCORE_COL = 'C'
LEFT_POINTS_COL = 'D'
RIGHT_POINTS_COL = 'F'
RIGHT_SCORE_COL = 'G'


MAX_LATE_VALUE = 1000
ABSENT_VALUE = 2000

ROUND1_LATE_TH = 10
ROUND2_LATE_TH = 30
MAX_SELECTION_RUN = 1000

GROUPS = ["Girls", "Group A", "Group B", "Group C", "Group D", "Women & Boys"]

COURTS_COUNT = 5
COURT_ALLOCATIONS = {1: ["Group A"],
                     2: ["Group B"],
                     3: ["Group C"],
                     4: ["Group D"],
                     5: ["Girls", "Women & Boys"]}

def parse_argument():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx_file",
            help="xlsx file with attendance and other place holder sheets")
    parser.add_argument("-c",
            "--run_count",
            type=int,
            default=MAX_SELECTION_RUN,
            help="All first \"a\" rounds to have successive turns for teams")
    parser.add_argument("-a",
            "--allow_successive",
            type=int,
            default=0,
            help="All first \"a\" rounds to have successive turns for teams")
    return parser.parse_args()

def retrive_teams(wb):
    ws = wb[ATTENDANCE_SHEET]
    teams = {}
    groups = {}

    for group in GROUPS:
        groups[group] = []

    for row in range(2, ws.max_row+1):
        group = ws.cell(row, 1).value
        team = ws.cell(row, 2).value
        player1 = ws.cell(row, 3).value
        attend1 = ws.cell(row, 4).value
        player2 = ws.cell(row, 5).value
        attend2 = ws.cell(row, 6).value
        if group and team and player1 and player2:
            teams[team] = {}
            teams[team]['rounds'] = []
            if attend1 != None:
                a1 = int(float(attend1))
            else:
                a1 = MAX_LATE_VALUE

            if attend2 != None:
                a2 = int(float(attend2))
            else:
                a2 = MAX_LATE_VALUE

            if a1 < 0:
                print(f'Error in sheet {ATTENDANCE_SHEET} at row {row}, invalid attendance value "{attend1}"')
                exit(1)

            if a2 < 0:
                print(f'Error in sheet {ATTENDANCE_SHEET} at row {row}, invalid attendance value "{attend2}"')
                exit(1)

            if a1 == 0 or a2 == 0:
                teams[team]['ready'] = ABSENT_VALUE
            else:
                teams[team]['ready'] = max(a1, a2)

            teams[team]['players'] = f'{player1} & {player2}'

            if group not in GROUPS:
                print(f'Error in sheet {ATTENDANCE_SHEET} at row {row}, unknown group "{group}"')
                exit(1)

            groups[group].append(team)
            teams[team]['group'] = group

    return teams, groups

def choose_matches(allow_successive, round_count, teams, quali_matches):
    rounds = [None for i in range(COURTS_COUNT)]
    available_rounds = []
    selected_teams = []
    choosen_count = 0

    for [left, right] in quali_matches:
        highest_ready = max(teams[left]['ready'], teams[right]['ready'])
        lowest_wait_time = min(teams[left]['wait_time'], teams[right]['wait_time'])
        entry = [left, right, highest_ready, lowest_wait_time]
        available_rounds.append(entry)

    available_rounds.sort(key=lambda k: (-1 * k[2], lowest_wait_time), reverse=True)

    for court in range(1, COURTS_COUNT+1):
        for i in range(len(available_rounds)):
            left = available_rounds[i][0]
            right = available_rounds[i][1]
            #print(f'{i}: avail {left} {right} {available_rounds[i][2]} {available_rounds[i][3]}')

            if left in selected_teams or right in selected_teams:
                continue

            if teams[left]['ready'] == ABSENT_VALUE or teams[right]['ready'] == ABSENT_VALUE:
                continue

            if not allow_successive and available_rounds[i][3] == 0:
                continue

            if teams[left]['group'] in COURT_ALLOCATIONS[court]:
                rounds[court-1] = [left, right]
                selected_teams.append(left)
                selected_teams.append(right)
                break

    return rounds

def update_chosen_teams(teams, quali_matches, rounds):

    selected_teams = []
    for match in rounds:
        if match:
            teams[match[0]]['wait_time'] = 0
            teams[match[1]]['wait_time'] = 0

            selected_teams.append(match[0])
            selected_teams.append(match[1])

            quali_matches.remove(match)

    for team in teams:
        if team not in selected_teams:
            teams[team]['wait_time'] += 1

def update_row(ws, row, iterable):
   for cell_row in ws.iter_rows(min_row=row, max_col=len(iterable), max_row=row):
       for i, cell in enumerate(cell_row):
           cell.value = iterable[i]

def update_schedule(wb, teams, quali_rounds):

    ws = wb[SCHEDULE_SHEET]
    for row in range(1, ws.max_row+1):
        for col in range(1, ws.max_column+1):
            ws.cell(row, col).value = None

    row = 1
    schedule_row = [""]
    for i in range(COURTS_COUNT):
        schedule_row.append(f'Court - {i+1}')
        schedule_row.append("")
        schedule_row.append("")
    update_row(ws, row, schedule_row)
    row += 1

    for seq, rounds in enumerate(quali_rounds):
        schedule_row = [""]
        for court, match in enumerate(rounds):
            if match:
                left = match[0]
                right = match[1]
                teams[left]['rounds'].append((seq+1, court+1))
                teams[right]['rounds'].append((seq+1, court+1))
                schedule_row.append(left)
                schedule_row.append(right)
                schedule_row.append("")
            else:
                schedule_row.append("")
                schedule_row.append("")
                schedule_row.append("")
        update_row(ws, row, schedule_row)
        row += 1

    INDIVIDUAL_TEAM_TABLE_ROW = 16

    row = INDIVIDUAL_TEAM_TABLE_ROW
    for col in range(3, len(quali_rounds) + 3):
        ws.cell(row, col, f'Round - {col-2}')

    row += 1
    for team in teams:
        ws.cell(row, 1, teams[team]['players'])
        ws.cell(row, 2, team)
        rounds = teams[team]['rounds']
        for (seq, court) in rounds:
            ws.cell(row, 2+seq, f'Court - {court}')
        row += 1

    for col in range(3, len(quali_rounds) + 3):
        ws.cell(row, col, f'Round - {col-2}')


def update_score_sheet(wb, groups, teams, quali_rounds):
    ws = wb[SCORE_SHEET]
    ws.delete_rows(1, ws.max_row)
    title = ["players1", "team1", "score", "points", "", "points", "score", "team2", "players2"]
    ws.append(title)
    row = 2

    for team in teams:
        teams[team]['score_rows'] = []

    for seq, rounds in enumerate(quali_rounds):
        ws.append([])
        row += 1
        ws.append([f'Round - {seq+1}'])
        row += 1
        for match in rounds:
            row_entries = ["" for x in title]
            if match:
                left = match[0]
                right = match[1]
                row_entries[0] = teams[left]['players']
                row_entries[1] = left
                left_points_eqn = f'=if({LEFT_SCORE_COL}{row} > {RIGHT_SCORE_COL}{row},{WINNER_POINTS},"")'
                row_entries[3] = left_points_eqn
                right_points_eqn = f'=if({LEFT_SCORE_COL}{row} < {RIGHT_SCORE_COL}{row},{WINNER_POINTS},"")'
                row_entries[5] = right_points_eqn
                row_entries[7] = right
                row_entries[8] = teams[right]['players']
                teams[left]['score_rows'].append(('left', row))
                teams[right]['score_rows'].append(('right', row))
                ws.append(row_entries)
                row += 1
    ws.append([])
    ws.append([])

    for team in teams:
        if not teams[team]['score_rows']:
            points_sum_eqn = ''
            score_sum_eqn = ''
        else:
            points_sum_eqn = '=sum('
            score_sum_eqn = '=sum('
            for (side, row) in teams[team]['score_rows']:
                if side == 'left':
                    points_sum_eqn += f'{LEFT_POINTS_COL}{row}, '
                    score_sum_eqn += f'{LEFT_SCORE_COL}{row}, '
                elif side == 'right':
                    points_sum_eqn += f'{RIGHT_POINTS_COL}{row}, '
                    score_sum_eqn += f'{RIGHT_SCORE_COL}{row}, '
            points_sum_eqn = points_sum_eqn[:-2] + ')'
            score_sum_eqn = score_sum_eqn[:-2] + ')'

        teams[team]['points_sum_eqn'] = points_sum_eqn
        teams[team]['score_sum_eqn'] = score_sum_eqn

    for group in groups:
        ws.append([group])
        for team in groups[group]:
            row_entries = ["" for i in range(7)]
            if teams[team]['ready'] != ABSENT_VALUE:
                row_entries[0] = teams[team]['players']
                row_entries[1] = team
                row_entries[2] = teams[team]['points_sum_eqn']
                row_entries[6] = teams[team]['score_sum_eqn']
            ws.append(row_entries)
        ws.append([])

    return

def run_quali_rounds(successive_rounds, groups, teams):

    quali_matches = []

    for team in teams:
        teams[team]['wait_time'] = 1

    for group in groups:
        matches = combinations(groups[group], 2)
        for match in matches:
            (left, right) = match
            if teams[left]['ready'] != ABSENT_VALUE and teams[right]['ready'] != ABSENT_VALUE:
                quali_matches.append(list(match))

    random.shuffle(quali_matches)

    quali_rounds = []
    round_count = 0
    while (quali_matches):
        round_count += 1
        if round_count <= successive_rounds:
            allow_successive = True
        else:
            allow_successive = False
        rounds = choose_matches(allow_successive, round_count, teams, quali_matches)
        update_chosen_teams(teams, quali_matches, rounds)
        quali_rounds.append(rounds)

    crt_rnd_cnts = [0 for i in range(COURTS_COUNT)]

    for n, rounds in enumerate(reversed(quali_rounds)):
        for i in range(COURTS_COUNT):
            if not crt_rnd_cnts[i] and rounds[i]:
                crt_rnd_cnts[i] = round_count - n

    total_round_count = sum(crt_rnd_cnts)

    roun1_violation = []
    roun2_violation = []
    for seq, rounds in enumerate(quali_rounds):
        for court, match in enumerate(rounds):
            if match:
                left = match[0]
                right = match[1]
                if seq == 0:
                    if teams[left]['ready'] > ROUND1_LATE_TH:
                        roun1_violation.append(left)
                    if teams[right]['ready'] > ROUND1_LATE_TH:
                        roun1_violation.append(right)
                elif seq == 1:
                    if teams[left]['ready'] > ROUND2_LATE_TH:
                        roun2_violation.append(left)
                    if teams[right]['ready'] > ROUND2_LATE_TH:
                        roun2_violation.append(right)
    return total_round_count, roun1_violation, roun2_violation, quali_rounds

if __name__ == "__main__":

    args =  parse_argument()

    excel_file = args.xlsx_file

    wb = openpyxl.load_workbook(filename=excel_file)
    teams, groups = retrive_teams(wb)

    best_round_count = 100
    best_round1_violations = [None for i in range(100)]
    best_round2_violations = [None for i in range(100)]
    best_rounds = None
    best_rounds = None

    for run in range(args.run_count):
        round_count, roun1_violation, roun2_violation, quali_rounds = run_quali_rounds(args.allow_successive, groups, teams)
        #print(f'round_count {round_count}, roun1_violation {len(roun1_violation)}, roun2_violation {len(roun2_violation)}')
        if round_count > best_round_count:
            continue
        elif round_count < best_round_count:
            best_round_count = round_count
            best_rounds = quali_rounds
            best_round1_violations = roun1_violation
            best_round2_violations = roun2_violation
        else:
            if len(roun1_violation) > len(best_round1_violations):
                continue
            elif len(roun1_violation) < len(best_round1_violations):
                best_rounds = quali_rounds
                best_round1_violations = roun1_violation
                best_round2_violations = roun2_violation
            else:
                if len(roun2_violation) > len(best_round2_violations):
                    continue
                elif len(roun2_violation) < len(best_round2_violations):
                    best_rounds = quali_rounds
                    best_round1_violations = roun1_violation
                    best_round2_violations = roun2_violation

    print(f'Total rounds {best_round_count}')
    if best_round1_violations:
        print(f'Round1 ready violations:')
        for team in best_round1_violations:
            print(f'  {team} {teams[team]["ready"]}')
    if best_round2_violations:
        print(f'Round2 ready violations:')
        for team in best_round2_violations:
            print(f'  {team} {teams[team]["ready"]}')

    for seq, rounds in enumerate(best_rounds):
        for court, match in enumerate(rounds):
            if match:
                left = match[0]
                right = match[1]
                print(f'({left:6} {right:6}), ', end='')
            else:
                print(f'({"":6} {"":6}), ', end='')
        print()


    update_schedule(wb, teams, best_rounds)

    for team in teams:
        if teams[team]['ready'] == ABSENT_VALUE:
            continue
        rounds = teams[team]['rounds']
        gaps = []
        for i in range(1, len(rounds)):
            gaps.append(rounds[i][0] - rounds[i-1][0])

        print(f'{team}: min {min(gaps)} max {max(gaps)} rounds {" ".join([str(x) for x in rounds])}')
    update_score_sheet(wb, groups, teams, best_rounds)
    wb.save(excel_file)
